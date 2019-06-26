import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
#Choose a working directory
os.chdir(r"C:\Users\TEF\Documents\Python_Demo")

def txt_parser(DataSource):
    #首先选择一个工作目录创建一个Excel文件：Result_Model.xlsx
    path = r"C:\Users\TEF\Desktop\result_output\Result_Model.xlsx"
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    
    #df1
    df1_n1 = Result.index('KPIStart')
    df1_n2 = Result.index('KPIEnd')
    df1_tmp1 = Result[(df1_n1+2):df1_n2]
    df1_tmp2 = pd.DataFrame(np.array(df1_tmp1),columns = ['ProductType'])
    df1 = pd.DataFrame(df1_tmp2.ProductType.str.split('/').tolist(),columns = ['KPI','Value'])
    a = ['Value']
    #将'Value'列对应的数据转变成float类型
    df1[a] = df1[a].astype(str).astype(float).round()
    df1.to_excel(writer, sheet_name = 'KPI',merge_cells=False)
    
    #df2
    df2_n1 = Result.index('ShiftPlan1Start')
    df2_n2 = Result.index('ShiftPlan1End')
    df2_tmp1 = Result[(df2_n1+2):df2_n2]
    df2_tmp2 = pd.DataFrame(np.array(df2_tmp1),columns = ['ProductType'])
    df2 = pd.DataFrame(df2_tmp2.ProductType.str.split('/').tolist(),columns = ['Operation','MAE_Qty','ShiftModel','WD','HC'])
    a = ['MAE_Qty','WD','HC']
    df2[a] = df2[a].astype(str).astype(float).round()
    df2.to_excel(writer, sheet_name = 'ShiftPlan1',merge_cells=False)
    writer.save()
    writer.close()
  
 if __name__ == "__main__":
    with open("originalData.txt") as f:
      content = f.readlines()
    #remove whitespace characters like `\n` at the end of each line
    datasource= [x.strip() for x in content]
    txt_parser(datasource)
    
    
